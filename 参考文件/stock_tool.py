import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
import json
import os
import pandas as pd
import efinance as ef
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 数据保存文件
DATA_FILE = "stocks.json"

class StockApp:
    def __init__(self, root):
        self.root = root
        self.root.title("股票数据追踪器 (Python版)")
        self.root.geometry("800x600")

        # 数据存储结构: list of dict {'code': '301418', 'name': '协昌科技', 'count': 1000}
        self.stocks = self.load_data()

        # UI 布局
        self.setup_ui()
        self.refresh_table_list()

    def setup_ui(self):
        # 顶部控制栏
        frame_controls = tk.Frame(self.root, pady=10)
        frame_controls.pack(fill=tk.X, padx=10)

        tk.Label(frame_controls, text="股票代码:").pack(side=tk.LEFT)
        self.entry_code = tk.Entry(frame_controls, width=10)
        self.entry_code.pack(side=tk.LEFT, padx=5)

        btn_add = tk.Button(frame_controls, text="添加股票", command=self.add_stock, bg="#007bff", fg="white")
        btn_add.pack(side=tk.LEFT, padx=5)

        tk.Label(frame_controls, text=" | ").pack(side=tk.LEFT)

        btn_refresh = tk.Button(frame_controls, text="刷新现价", command=self.refresh_prices, bg="#17a2b8", fg="white")
        btn_refresh.pack(side=tk.LEFT, padx=5)

        tk.Label(frame_controls, text=" | ").pack(side=tk.LEFT)

        tk.Label(frame_controls, text="导出天数:").pack(side=tk.LEFT)
        self.entry_days = tk.Entry(frame_controls, width=5)
        self.entry_days.insert(0, "30")
        self.entry_days.pack(side=tk.LEFT, padx=5)

        btn_export = tk.Button(frame_controls, text="导出Excel", command=self.export_excel, bg="#28a745", fg="white")
        btn_export.pack(side=tk.LEFT, padx=5)

        # 股票列表 (Treeview)
        frame_table = tk.Frame(self.root)
        frame_table.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        columns = ("code", "name", "count", "latest_price", "total_val")
        self.tree = ttk.Treeview(frame_table, columns=columns, show="headings")
        
        self.tree.heading("code", text="代码")
        self.tree.heading("name", text="名称")
        self.tree.heading("count", text="持仓数量 (双击修改)")
        self.tree.heading("latest_price", text="参考现价")
        self.tree.heading("total_val", text="参考市值")

        self.tree.column("code", width=100, anchor=tk.CENTER)
        self.tree.column("name", width=150, anchor=tk.CENTER)
        self.tree.column("count", width=120, anchor=tk.CENTER)
        self.tree.column("latest_price", width=100, anchor=tk.E)
        self.tree.column("total_val", width=120, anchor=tk.E)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # 滚动条
        scrollbar = ttk.Scrollbar(frame_table, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 绑定双击事件 (修改持仓)
        self.tree.bind("<Double-1>", self.on_double_click)
        # 绑定右键事件 (删除)
        self.tree.bind("<Button-3>", self.show_context_menu)

        # 底部信息
        self.lbl_status = tk.Label(self.root, text="就绪", bd=1, relief=tk.SUNKEN, anchor=tk.W)
        self.lbl_status.pack(side=tk.BOTTOM, fill=tk.X)

        # 右键菜单
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="删除股票", command=self.delete_selected)

    def load_data(self):
        if os.path.exists(DATA_FILE):
            try:
                with open(DATA_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except:
                return []
        return []

    def save_data(self):
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(self.stocks, f, ensure_ascii=False, indent=2)

    def add_stock(self):
        code = self.entry_code.get().strip()
        if not code or len(code) != 6:
            messagebox.showerror("错误", "请输入6位股票代码")
            return

        # 检查是否已存在
        for s in self.stocks:
            if s['code'] == code:
                messagebox.showwarning("提示", "该股票已在列表中")
                return

        self.lbl_status.config(text="正在获取股票名称...")
        self.root.update()

        try:
            # 尝试获取最近历史数据来验证代码并获取名称
            df = ef.stock.get_quote_history(code)
            if df is None or df.empty:
                raise Exception("未找到股票数据")
            
            # 提取名称 (通常第一行包含名称)
            name = df.iloc[0]['股票名称']
            latest_close = df.iloc[-1]['收盘']

            new_stock = {
                'code': code,
                'name': name,
                'count': 0,
                'latest_price': float(latest_close)
            }
            self.stocks.append(new_stock)
            self.save_data()
            self.refresh_table_list()
            self.entry_code.delete(0, tk.END)
            self.lbl_status.config(text=f"成功添加: {name}")

        except Exception as e:
            messagebox.showerror("错误", f"获取股票信息失败: {str(e)}")
            self.lbl_status.config(text="就绪")

    def refresh_prices(self):
        if not self.stocks:
            return

        self.lbl_status.config(text="正在刷新最新价格...")
        self.root.update()

        try:
            codes = [s['code'] for s in self.stocks]
            # 批量获取最新行情
            df = ef.stock.get_latest_quote(codes)
            
            # 创建代码到价格的映射
            price_map = {}
            if df is not None and not df.empty:
                for _, row in df.iterrows():
                    # 确保代码格式一致 (efinance返回的是字符串)
                    c = str(row['代码'])
                    p = row['最新价']
                    # 有些情况可能返回 '-' (停牌等)
                    try:
                        price_map[c] = float(p)
                    except:
                        pass
            
            # 更新本地数据
            updated_count = 0
            for s in self.stocks:
                c = s['code']
                if c in price_map:
                    s['latest_price'] = price_map[c]
                    updated_count += 1
            
            self.save_data()
            self.refresh_table_list()
            self.lbl_status.config(text=f"刷新完成，已更新 {updated_count} 只股票价格")
            
        except Exception as e:
            messagebox.showerror("刷新失败", f"刷新价格失败: {str(e)}")
            self.lbl_status.config(text="刷新失败")

    def refresh_table_list(self):
        # 清空列表
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        total_asset = 0
        for stock in self.stocks:
            price = stock.get('latest_price', 0)
            count = stock.get('count', 0)
            val = price * count
            total_asset += val
            self.tree.insert("", tk.END, values=(stock['code'], stock['name'], count, price, f"{val:.2f}"))
        
        self.lbl_status.config(text=f"合计参考市值: {total_asset:.2f} 元 (价格仅供参考，请以导出数据为准)")

    def on_double_click(self, event):
        item_id = self.tree.selection()[0]
        item = self.tree.item(item_id)
        code = item['values'][0]
        current_count = item['values'][2]

        new_count = simpledialog.askfloat("修改持仓", f"请输入 {code} 的持仓数量:", initialvalue=current_count)
        if new_count is not None:
            for s in self.stocks:
                if s['code'] == str(code): # treeview values might be converted to int if numeric
                    s['count'] = new_count
                    break
                # Handle possible string/int mismatch
                if str(s['code']) == str(code):
                    s['count'] = new_count
                    break
            
            self.save_data()
            self.refresh_table_list()

    def show_context_menu(self, event):
        try:
            self.tree.selection_set(self.tree.identify_row(event.y))
            self.context_menu.post(event.x_root, event.y_root)
        except:
            pass

    def delete_selected(self):
        selected_item = self.tree.selection()
        if not selected_item:
            return
        item = self.tree.item(selected_item[0])
        code = str(item['values'][0])

        if messagebox.askyesno("确认", f"确定删除股票 {code} 吗?"):
            self.stocks = [s for s in self.stocks if str(s['code']) != code]
            self.save_data()
            self.refresh_table_list()

    def export_excel(self):
        if not self.stocks:
            messagebox.showwarning("提示", "没有股票可导出")
            return

        try:
            days = int(self.entry_days.get())
        except:
            days = 30

        self.lbl_status.config(text="正在批量获取历史数据，请稍候...")
        self.root.update()

        all_dfs = []
        codes = [s['code'] for s in self.stocks]

        # 颜色配置 (浅色背景)
        bg_colors = ["D9EAD3", "FCE5CD", "C9DAF8", "EAD1DC", "FFF2CC", "D0E0E3"]
        
        try:
            # 使用 efinance 批量获取数据可能会更快，但为了准确控制，我们循环获取
            # 或者使用 ef.stock.get_quote_history(codes) 如果支持列表 (efinance supports dict or list usually)
            # 为了稳妥，我们一个个获取并合并
            
            # 创建一个基准日期DataFrame
            merged_df = pd.DataFrame()

            for i, stock in enumerate(self.stocks):
                code = stock['code']
                name = stock['name']
                count = stock['count']

                # 获取历史K线
                df = ef.stock.get_quote_history(code)
                # ef 返回列: 股票名称, 股票代码, 日期, 开盘, 收盘, 最高, 最低, 成交量...
                
                # 筛选最近N天 (注意: efinance返回是按时间正序)
                df = df.tail(days).copy()
                
                # 保留需要的列并重命名，防止合并冲突
                # 我们需要: 日期 (作为Key), 收盘 (作为价格)
                # 持仓数是固定的
                
                # 在df中添加持仓和总价列
                df['count'] = count
                df['total'] = df['收盘'] * count
                
                # 重命名列以便区分
                # 格式: Date, {code}_count, {code}_price, {code}_total
                part_df = df[['日期', 'count', '收盘', 'total']].copy()
                part_df.rename(columns={
                    'count': f'{code}_count', 
                    '收盘': f'{code}_price', 
                    'total': f'{code}_total'
                }, inplace=True)
                
                if merged_df.empty:
                    merged_df = part_df
                else:
                    # Outer Merge 保证日期对齐，即使某股票某天停牌
                    merged_df = pd.merge(merged_df, part_df, on='日期', how='outer')

            # 按日期降序排列
            merged_df['日期'] = pd.to_datetime(merged_df['日期'])
            merged_df.sort_values('日期', ascending=False, inplace=True)
            merged_df['日期'] = merged_df['日期'].dt.strftime('%Y-%m-%d')
            
            # --- OpenPyXL 导出美化 ---
            wb = Workbook()
            ws = wb.active
            ws.title = "横向对比"

            # 样式定义
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin'))
            align_center = Alignment(horizontal='center', vertical='center')
            font_bold = Font(bold=True)
            font_header = Font(bold=True, size=12)

            # 写入表头 Row 1 (股票名称) & Row 2 (子标题)
            # A1: 日期
            ws.merge_cells('A1:A2')
            cell_a1 = ws['A1']
            cell_a1.value = "日期"
            cell_a1.alignment = align_center
            cell_a1.border = thin_border
            ws.column_dimensions['A'].width = 12

            current_col = 2 # B列开始
            
            for i, stock in enumerate(self.stocks):
                code = stock['code']
                name = stock['name']
                
                # 表头背景色
                color_hex = bg_colors[i % len(bg_colors)]
                fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                
                # 写入 Row 1: 股票名称 (合并3列)
                # 比如 B1:D1
                start_col_letter = get_column_letter(current_col)
                end_col_letter = get_column_letter(current_col + 2)
                merge_range = f"{start_col_letter}1:{end_col_letter}1"
                
                ws.merge_cells(merge_range)
                top_cell = ws[f"{start_col_letter}1"]
                top_cell.value = f"{name} ({code})"
                top_cell.fill = fill
                top_cell.font = font_header
                top_cell.alignment = align_center
                top_cell.border = thin_border
                
                # 给合并单元格的所有格子加边框/背景(可选，openpyxl有时候只渲染左上角)
                # 简单起见，只处理左上角，通常足够。若需完美边框需遍历范围。

                # 写入 Row 2: 股数 | 价格 | 总价
                sub_headers = ["股数", "价格", "总价"]
                for idx, title in enumerate(sub_headers):
                    c_letter = get_column_letter(current_col + idx)
                    cell = ws[f"{c_letter}2"]
                    cell.value = title
                    cell.font = font_bold
                    cell.alignment = align_center
                    cell.border = thin_border
                    # 设置浅灰背景
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    
                    # 设置列宽
                    width = 12 if idx == 2 else 10
                    ws.column_dimensions[c_letter].width = width

                current_col += 3
            
            # 写入数据
            # merged_df 的列现在是: 日期, 301418_count, 301418_price, 301418_total, 600519_count...
            # 我们需要按照 self.stocks 的顺序读取列
            
            row_idx = 3
            for _, row_data in merged_df.iterrows():
                ws.cell(row=row_idx, column=1, value=row_data['日期']).border = thin_border
                
                current_col = 2
                for stock in self.stocks:
                    code = stock['code']
                    # 获取该行该股票的数据
                    c_val = row_data.get(f'{code}_count')
                    p_val = row_data.get(f'{code}_price')
                    t_val = row_data.get(f'{code}_total')
                    
                    # 检查是否为空 (NaN)
                    if pd.isna(p_val):
                        # 数据填充为空
                        ws.cell(row=row_idx, column=current_col).border = thin_border
                        ws.cell(row=row_idx, column=current_col+1).border = thin_border
                        ws.cell(row=row_idx, column=current_col+2).border = thin_border
                    else:
                        c_cell = ws.cell(row=row_idx, column=current_col, value=c_val)
                        p_cell = ws.cell(row=row_idx, column=current_col+1, value=p_val)
                        t_cell = ws.cell(row=row_idx, column=current_col+2, value=t_val)
                        
                        c_cell.border = thin_border
                        p_cell.border = thin_border
                        t_cell.border = thin_border

                    current_col += 3
                row_idx += 1

            filename = f"股票横向对比_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            self.lbl_status.config(text=f"导出成功: {filename}")
            messagebox.showinfo("成功", f"文件已保存为:\n{filename}")
            
            # 顺便打开文件
            try:
                os.startfile(filename)
            except:
                pass

        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("导出失败", f"发生错误: {str(e)}")
            self.lbl_status.config(text="导出失败")

if __name__ == "__main__":
    root = tk.Tk()
    app = StockApp(root)
    root.mainloop()
